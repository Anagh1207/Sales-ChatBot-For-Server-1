"""
Load Excel workbooks into PostgreSQL using pandas + openpyxl engine.
"""
from __future__ import annotations

import logging
import os

import pandas as pd
from sqlalchemy.engine import Engine

from app.core.config import settings
from app.services.schema_infer import recreate_table_from_dataframe
from app.services.table_metadata import invalidate_table_cache

logger = logging.getLogger(__name__)


def _read_excel(path: str) -> pd.DataFrame:
    if not os.path.isfile(path):
        raise FileNotFoundError(f"Excel file not found: {path}")
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    sheet_name = wb.sheetnames[0]
    if "Export" in wb.sheetnames:
        sheet_name = "Export"
    df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
    
    # Coerce likely date columns to datetime so PostgreSQL stores typed timestamps.
    from app.services.table_metadata import normalize_header

    for col in list(df.columns):
        if "date" in normalize_header(str(col)):
            df[col] = pd.to_datetime(df[col], errors="coerce")
    return df


def ingest_sales(engine: Engine, path: str | None = None) -> int:
    """Load Sales.xlsx into sales_data. Returns row count."""
    p = path or settings.sales_excel_path
    df = _read_excel(p)
    recreate_table_from_dataframe(engine, df, "sales_data")
    invalidate_table_cache("sales_data")  # force fresh reflection after schema recreate
    return len(df)


def ingest_timesheets(engine: Engine, path: str | None = None) -> int:
    """Load Timesheet.xlsx into timesheet_data. Returns row count (disabled in single-table mode)."""
    return 0


def ingest_all(engine: Engine) -> dict[str, int]:
    """Ingest both workbooks (timesheets disabled)."""
    return {
        "sales_data": ingest_sales(engine),
        "timesheet_data": ingest_timesheets(engine),
    }

